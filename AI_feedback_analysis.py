import streamlit as st
import pandas as pd
import sqlite3
import google.generativeai as genai
from datetime import datetime
from openpyxl import Workbook
import os
import io

# GEMINI CONFIGURATION

GEMINI_API_KEY = "API"  

genai.configure(api_key=GEMINI_API_KEY)

model = genai.GenerativeModel("gemini-2.5-flash")


# DATABASE FUNCTIONS

DB_NAME = "feedback_analysis.db"


def create_database():
    conn = sqlite3.connect(DB_NAME)
    cursor = conn.cursor()

    cursor.execute("""
    CREATE TABLE IF NOT EXISTS FeedbackAnalysis(
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        feedback TEXT,
        sentiment TEXT,
        category TEXT,
        summary TEXT,
        analyzed_at TEXT
    )
    """)

    conn.commit()
    conn.close()


def insert_feedback(feedback, sentiment, category, summary):
    conn = sqlite3.connect(DB_NAME)
    cursor = conn.cursor()

    cursor.execute("""
    INSERT INTO FeedbackAnalysis
    (
        feedback,
        sentiment,
        category,
        summary,
        analyzed_at
    )
    VALUES (?, ?, ?, ?, ?)
    """,
    (
        feedback,
        sentiment,
        category,
        summary,
        datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ))

    conn.commit()
    conn.close()


def get_all_feedback():
    conn = sqlite3.connect(DB_NAME)

    df = pd.read_sql_query(
        "SELECT * FROM FeedbackAnalysis",
        conn
    )

    conn.close()

    return df


def search_feedback(sentiment):
    conn = sqlite3.connect(DB_NAME)

    query = """
    SELECT *
    FROM FeedbackAnalysis
    WHERE sentiment = ?
    """

    df = pd.read_sql_query(
        query,
        conn,
        params=(sentiment,)
    )

    conn.close()

    return df

def delete_feedback(feedback_id):
    conn = sqlite3.connect(DB_NAME)
    cursor = conn.cursor()

    cursor.execute(
        "DELETE FROM FeedbackAnalysis WHERE id = ?",
        (feedback_id,)
    )

    conn.commit()
    rows_deleted = cursor.rowcount
    conn.close()

    return rows_deleted


# GEMINI ANALYSIS

def analyze_feedback(feedback):

    prompt = f"""
Analyze the following customer feedback.

Feedback:
{feedback}

Return in exactly this format:

Sentiment: Positive/Negative/Neutral
Category: Product/Delivery/Customer Support/Pricing/Other
Summary: One sentence summary
"""

    try:
        response = model.generate_content(prompt)

        result = response.text.strip()

        sentiment = "Unknown"
        category = "Other"
        summary = ""

        lines = result.split("\n")

        for line in lines:

            if line.startswith("Sentiment"):
                sentiment = line.split(":", 1)[1].strip()

            elif line.startswith("Category"):
                category = line.split(":", 1)[1].strip()

            elif line.startswith("Summary"):
                summary = line.split(":", 1)[1].strip()

        return sentiment, category, summary

    except Exception as e:
        return "Error", "Error", str(e)


# EXPORT REPORT

def export_report():

    df = get_all_feedback()

    if df.empty:
        return None

    file_name = "feedback_analysis_report.xlsx"

    wb = Workbook()

    ws1 = wb.active
    ws1.title = "Analysis Results"

    headers = [
        "ID",
        "Feedback",
        "Sentiment",
        "Category",
        "Summary",
        "Analyzed At"
    ]

    ws1.append(headers)

    for row in df.values.tolist():
        ws1.append(row)

    # Summary Sheet
    ws2 = wb.create_sheet("Summary")

    total = len(df)

    positive = len(
        df[df["sentiment"].str.lower() == "positive"]
    )

    negative = len(
        df[df["sentiment"].str.lower() == "negative"]
    )

    neutral = len(
        df[df["sentiment"].str.lower() == "neutral"]
    )

    ws2.append(["Metric", "Value"])
    ws2.append(["Total Feedback", total])
    ws2.append(["Positive", positive])
    ws2.append(["Negative", negative])
    ws2.append(["Neutral", neutral])

    wb.save(file_name)

    return file_name


# INITIALIZATION

create_database()

st.set_page_config(
    page_title="AI Feedback Analysis Dashboard",
    layout="wide"
)

st.title(" AI Feedback Analysis Dashboard")

# SIDEBAR MENU

menu = st.sidebar.radio(
    "Navigation",
    [
        "Upload & Analyze",
        "Dashboard",
        "Search Feedback",
        "Delete Feedback",
        "Export Report"
    ]
)

# UPLOAD & ANALYZE

if menu == "Upload & Analyze":

    st.header("Add Feedback")

    input_mode = st.radio(
        "How would you like to add feedback?",
        ["Write Manually", "Upload File"],
        horizontal=True,
        key="input_mode_selector"
    )

    # MANUAL ENTRY MODE
    if input_mode == "Write Manually":

        feedback_text = st.text_area(
            "Enter customer feedback",
            height=150,
            placeholder="Type or paste the customer feedback here..."
        )

        if st.button("Analyze Feedback"):

            if not feedback_text.strip():
                st.warning("Please enter some feedback before analyzing.")

            else:
                with st.spinner("Analyzing feedback..."):

                    sentiment, category, summary = analyze_feedback(
                        feedback_text.strip()
                    )

                    insert_feedback(
                        feedback_text.strip(),
                        sentiment,
                        category,
                        summary
                    )

                st.success("Feedback analyzed and stored in database.")

                col1, col2, col3 = st.columns(3)

                col1.metric("Sentiment", sentiment)
                col2.metric("Category", category)

                col3.write("**Summary**")
                col3.write(summary)

                # EXPORT THIS SINGLE ENTRY

                wb = Workbook()
                ws = wb.active
                ws.title = "Feedback"

                ws.append([
                    "Feedback",
                    "Sentiment",
                    "Category",
                    "Summary",
                    "Analyzed At"
                ])

                ws.append([
                    feedback_text.strip(),
                    sentiment,
                    category,
                    summary,
                    datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                ])

                buffer = io.BytesIO()
                wb.save(buffer)
                buffer.seek(0)

                st.download_button(
                    label="Download This Entry",
                    data=buffer,
                    file_name="manual_feedback_entry.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key=f"download_{datetime.now().strftime('%Y%m%d%H%M%S%f')}"
                )

    # FILE UPLOAD MODE
    else:

        uploaded_file = st.file_uploader(
            "Upload feedback.xlsx",
            type=["xlsx"]
        )

        if uploaded_file:

            df = pd.read_excel(uploaded_file)

            st.subheader("Uploaded Data")
            st.dataframe(df)

            if st.button("Analyze Uploaded Feedback"):

                progress = st.progress(0)

                total_rows = len(df)

                for index, row in df.iterrows():

                    feedback = str(row["Feedback"])

                    sentiment, category, summary = analyze_feedback(
                        feedback
                    )

                    insert_feedback(
                        feedback,
                        sentiment,
                        category,
                        summary
                    )

                    progress.progress(
                        (index + 1) / total_rows
                    )

                st.success(
                    "Feedback analysis completed and stored in database."
                )

# DASHBOARD

elif menu == "Dashboard":

    st.header("Dashboard")

    df = get_all_feedback()

    if df.empty:
        st.warning("No data available.")
    else:

        total = len(df)

        positive = len(
            df[df["sentiment"].str.lower() == "positive"]
        )

        negative = len(
            df[df["sentiment"].str.lower() == "negative"]
        )

        neutral = len(
            df[df["sentiment"].str.lower() == "neutral"]
        )

        col1, col2, col3, col4 = st.columns(4)

        col1.metric("Total Feedback", total)
        col2.metric("Positive", positive)
        col3.metric("Negative", negative)
        col4.metric("Neutral", neutral)

        st.subheader("Sentiment Distribution")

        chart_data = pd.DataFrame({
            "Count": [
                positive,
                negative,
                neutral
            ]
        },
        index=[
            "Positive",
            "Negative",
            "Neutral"
        ])

        st.bar_chart(chart_data)

        st.subheader("All Feedback Records")

        st.dataframe(df)

# SEARCH FEEDBACK

elif menu == "Search Feedback":

    st.header("Search Feedback By Sentiment")

    sentiment = st.selectbox(
        "Select Sentiment",
        [
            "Positive",
            "Negative",
            "Neutral"
        ]
    )

    if st.button("Search"):

        result = search_feedback(sentiment)

        st.write(
            f"Found {len(result)} record(s)"
        )

        st.dataframe(result)


# DELETE FEEDBACK

elif menu == "Delete Feedback":

    st.header("Delete Feedback By ID")

    df = get_all_feedback()

    if df.empty:
        st.warning("No data available to delete.")
    else:
        st.dataframe(df)

        feedback_id = st.number_input(
            "Enter the ID of the feedback to delete",
            min_value=1,
            step=1
        )

        if st.button("Delete Feedback"):

            rows_deleted = delete_feedback(int(feedback_id))

            if rows_deleted > 0:
                st.success(f"Feedback with ID {int(feedback_id)} deleted.")
                st.rerun()
            else:
                st.warning(f"No feedback found with ID {int(feedback_id)}.")

# EXPORT REPORT
elif menu == "Export Report":

    st.header("Export Excel Report")

    if st.button("Generate Report"):

        file_name = export_report()

        if file_name:

            with open(file_name, "rb") as file:

                st.download_button(
                    label="Download Report",
                    data=file,
                    file_name=file_name,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="export_report_download"
                )

            st.success(
                "Report generated successfully."
            )

        else:
            st.warning(
                "No data available for export."
            )